import json
import time
from operator import itemgetter
import openpyxl
import requests
import re
json_str = []  # 储存所有json数据
data1 = []
list1 = []
list2 = []
for i in range(1, 100):
    time.sleep(1)
    response = requests.get(url='https://oj.qd.sdu.edu.cn/api/submit/list?pageNow='+str(i)+'&pageSize=20')
    page = response.content.decode()
    rs = json.loads(page)
    json_str.append(rs)
#    print(rs)
    temp = rs['data']['rows']
    for items in temp:
        keys = ['problemTitle', 'judgeScore']
        out = itemgetter(*keys)(items)
        data1.append(out)
        if items['judgeScore'] == 100:
            list1.append(items['problemTitle'])
        list2.append(items['problemTitle'])
count = []
dic1 = {}   # 通过量
dic2 = {}    # 提交量
for item in list1:
    count.append(list1.count(item))
dic1 = dict(zip(list1, count))
for item in list2:
    count.append(list2.count(item))
dic2 = dict(zip(list2, count))
dic1_sort = sorted(dict.items(dic1), key=lambda e: e[1], reverse=True)
dic2_sort = sorted(dict.items(dic2), key=lambda e: e[1], reverse=True)


print(dic1_sort)
print(dic2_sort)


# print(data1)
wb = openpyxl.Workbook()
ws1 = wb.active
ws1.title = 'sheet'
ws1.append(('problemTitle', 'judgeScore'))
for row in data1:
    ws1.append(row)
wb.save('C:\\Users\Lenovo\Desktop\\3.1.2.xls')

wb = openpyxl.Workbook()
ws2 = wb.active
ws2.title = 'sheet'
ws2.append(('problemTitle', 'ac'))
i = 0
for row in dic1_sort:
    if i < 10:
        ws2.append(row)
        i += 1
    else:
        break
wb.save('C:\\Users\Lenovo\Desktop\\3.1.2.1.xls')

wb = openpyxl.Workbook()
ws3 = wb.active
ws3.title = 'sheet'
ws3.append(('problemTitle', 'submit'))
j = 0
for row in dic2_sort:
    if j < 10:
        ws3.append(row)
        j += 1
    else:
        break
wb.save('C:\\Users\Lenovo\Desktop\\3.1.2.2.xls')

