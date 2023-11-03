import requests
import json
import time
from operator import itemgetter
import openpyxl
json_str = []
json_str2 = []
data1 = []
data2 = []
list1 = []
list2 = []
list_me = []

headers = {
     'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36',
     'cookie': 'SESSION=12312-5b13487d-412b-488b-9540-09a8cb75805f'
}

url = 'https://oj.qd.sdu.edu.cn/api/contest/query?contestId=286'
response = requests.get(url, headers=headers)
page = response.content.decode()
rs = json.loads(page)
json_str.append(rs)
temp = rs['data']['problems']
for items in temp:
    keys = ['problemCode',  'problemTitle', 'problemWeight', 'problemColor', 'acceptNum', 'submitNum', 'judgeScore']
    out = itemgetter(*keys)(items)
    data1.append(out)
    if items['judgeScore'] != 100:
        list_me.append(items['problemTitle'])
wb = openpyxl.Workbook()
ws1 = wb.active
ws1.title = 'sheet'
ws1.append(('problemCode',  'problemTitle', 'problemWeight', 'problemColor', 'acceptNum', 'submitNum', 'judgeScore'))
for row in data1:
    ws1.append(row)
wb.save('C:\\Users\Lenovo\Desktop\\3.2.1.xls')


for i in range(1, 4):
    time.sleep(1)
    response1 = requests.get(url='https://oj.qd.sdu.edu.cn/api/contest/listSubmission?pageNow='+str(i)+'&pageSize=20&contestId=286',headers=headers)
    page1 = response1.content.decode()
    rs1 = json.loads(page1)
    json_str2.append(rs1)
    temp = rs1['data']['rows']
    for items in temp:
        keys = ['submissionId',  'problemTitle', 'username', 'judgeScore', 'usedTime', 'usedMemory']
        out = itemgetter(*keys)(items)
        if items['judgeScore'] == 100:
            list1.append(items['username'])
        list2.append(items['username'])
        data2.append(out)
wb = openpyxl.Workbook()
ws2 = wb.active
ws2.title = 'sheet'
ws2.append(('submissionId',  'problemTitle', 'username', 'judgeScore', 'usedTime', 'usedMemory'))
for row in data2:
    ws2.append(row)
wb.save('C:\\Users\Lenovo\Desktop\\3.2.2.xls')

count1 = []
count2 = []

for item in list1:
    count1.append(list1.count(item))
li1 = list(zip(list1, count1))
for item in list2:
    count2.append(list2.count(item))
li2 = list(zip(list2, count2))

wb = openpyxl.Workbook()
ws3 = wb.active
ws3.title = 'sheet'
ws3.append(('username', 'acNum'))
for items in li1:
    ws3.append(items)
wb.save('C:\\Users\Lenovo\Desktop\\3.2.3.xls')

li4=[]
for ctc in range(1, 7):
    li4.append('OJ2023u010')
wb = openpyxl.Workbook()
ws4 = wb.active
ws4.title = 'sheet'
ws4.append(('username', 'subNum'))
for items in li2:
    ws4.append(items)
wb.save('C:\\Users\Lenovo\Desktop\\3.2.4.xls')

li3 = list(zip(li4, list_me))
wb = openpyxl.Workbook()
ws5 = wb.active
ws5.title = 'sheet'
ws5.append(('username', 'not_pass'))
for items in li3:
    ws5.append(items)
wb.save('C:\\Users\Lenovo\Desktop\\3.2.5.xls')